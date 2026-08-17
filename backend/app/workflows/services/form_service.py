"""Form service – form schema CRUD, bulk upload, and entity field sync.

Strict rules enforced here:
  - A form CANNOT be created without a pre-existing entity.
  - form_id is always provided by the caller (no auto-generation here).
  - Any new form fields are synced back to entity_fields via
    entity_service.sync_entity_fields_from_form.
  - Entity is the source of truth — forms extend it but cannot bypass it.
"""

from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Any

from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified

from app.models.entity import EntityDefinition, EntityRecord
from app.models.form import Form
from app.schemas.form import FormSchemaCreate, FormSchemaUpdate


# ── Helpers ────────────────────────────────────────────────────────────────────


def _require_entity_exists(db: Session, entity_name: str) -> EntityDefinition:
    """Return the EntityDefinition or raise ValueError.

    STRICT RULE: Forms cannot be created without a pre-existing entity.
    Auto-creation of entities from forms is explicitly forbidden.
    """
    entity = db.query(EntityDefinition).filter(EntityDefinition.name == entity_name).first()
    if not entity:
        raise ValueError(
            f"Entity '{entity_name}' does not exist. "
            "Create the entity first before creating a form that references it."
        )
    return entity


def _get_fields(form: Form) -> list[dict]:
    """Return the fields list from the form schema."""
    schema = form.schema or {}
    return schema.get("fields", [])


def _sync_fields_to_entity(db: Session, entity_def: EntityDefinition, form_schema: dict) -> None:
    """Sync form fields back to entity_fields (entity is source of truth)."""
    from .entity_service import sync_entity_fields_from_form
    form_fields = form_schema.get("fields", [])
    if form_fields:
        sync_entity_fields_from_form(db, entity_def, form_fields)


# ── CRUD ───────────────────────────────────────────────────────────────────────


def get_forms(db: Session, skip: int = 0, limit: int = 100) -> list[Form]:
    return db.query(Form).offset(skip).limit(limit).all()


def get_form_schema(db: Session, form_id: str) -> Form | None:
    return db.query(Form).filter(Form.form_id == form_id).first()


def create_form_schema(db: Session, form_in: FormSchemaCreate) -> Form:
    """Create a form schema.

    Enforces:
      1. Entity must already exist (no auto-creation).
      2. Form fields are synced back to entity_fields.
    """
    entity_def = _require_entity_exists(db, form_in.entity_name)

    # Sync any new fields from the form back to the entity
    if form_in.schema:
        _sync_fields_to_entity(db, entity_def, form_in.schema)

    db_form = Form(
        form_id=form_in.form_id,
        entity_name=form_in.entity_name,
        schema=form_in.schema,
    )
    db.add(db_form)
    db.commit()
    db.refresh(db_form)

    try:
        from app.services.dynamic_projection_service import (
            has_projection_table,
            sync_projection_schema,
        )
        if has_projection_table(db, entity_def.name):
            sync_projection_schema(db, entity_def.name)
    except Exception:
        pass  # projection sync failure must not block form save

    return db_form


def delete_form_schema(db: Session, form_id: str) -> bool:
    """Delete a form schema by form_id. Returns True if deleted, False if not found."""
    form = db.query(Form).filter(Form.form_id == form_id).first()
    if not form:
        return False
    db.delete(form)
    db.commit()
    return True


def update_form_schema(db: Session, form_id: str, form_in: FormSchemaUpdate) -> Form | None:
    """Update a form schema.

    Enforces:
      1. If entity_name changes, the new entity must already exist.
      2. Form fields are synced back to entity_fields on every update.
      3. Renamed fields are detected by comparing old vs new field IDs.
    """
    db_form = get_form_schema(db, form_id)
    if not db_form:
        return None

    if form_in.entity_name is not None:
        entity_def = _require_entity_exists(db, form_in.entity_name)
        db_form.entity_name = form_in.entity_name
    else:
        entity_def = _require_entity_exists(db, db_form.entity_name)

    if form_in.schema is not None:
        # Pass old form fields so rename detection can work
        old_fields = (db_form.schema or {}).get("fields", [])
        from .entity_service import sync_entity_fields_from_form
        new_fields = form_in.schema.get("fields", [])
        if new_fields:
            sync_entity_fields_from_form(db, entity_def, new_fields, old_form_fields=old_fields)
        db_form.schema = form_in.schema
        flag_modified(db_form, "schema")

    db.commit()
    db.refresh(db_form)

    # Sync projection schema — add columns for any new fields (never drops columns)
    try:
        from app.services.dynamic_projection_service import (
            has_projection_table,
            sync_projection_schema,
        )
        if has_projection_table(db, entity_def.name):
            sync_projection_schema(db, entity_def.name)
    except Exception:
        pass  # projection sync failure must not block form save

    return db_form


# ── Bulk Upload ────────────────────────────────────────────────────────────────


def _build_header_row(form: Form) -> list[str]:
    """Return column headers: asset_id first, then one column per form field."""
    fields = _get_fields(form)
    return ["asset_id"] + [f.get("label", f.get("id", "")) for f in fields]


def _build_hint_row(form: Form) -> list[str]:
    """Return a hints row describing each field (type + options)."""
    fields = _get_fields(form)
    hints = ["Asset / equipment identifier"]
    for f in fields:
        ftype = f.get("type", "text")
        options = f.get("options", [])
        required = " [required]" if f.get("required") else ""
        if options:
            hint = f"{ftype}{required} | options: {', '.join(str(o) for o in options)}"
        else:
            hint = f"{ftype}{required}"
        hints.append(hint)
    return hints


def generate_bulk_template_csv(form: Form) -> bytes:
    """Generate a CSV template with headers and a hints row."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_build_header_row(form))
    writer.writerow(_build_hint_row(form))
    blank = [""] * (len(_get_fields(form)) + 1)
    writer.writerow(blank)
    writer.writerow(blank)
    return buf.getvalue().encode("utf-8-sig")


def generate_bulk_template_xlsx(form: Form) -> bytes:
    """Generate an Excel template with styled headers and a hints row."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel export")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Upload"

    header_row = _build_header_row(form)
    hint_row = _build_hint_row(form)

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    hint_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    hint_font = Font(italic=True, color="374151")

    for col_idx, header in enumerate(header_row, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(18, len(header) + 4)

    for col_idx, hint in enumerate(hint_row, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hint)
        cell.font = hint_font
        cell.fill = hint_fill
        cell.alignment = Alignment(wrap_text=True)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_bulk_upload(form: Form, raw_bytes: bytes, filename: str) -> dict[str, Any]:
    """Parse a CSV or Excel file and return structured preview data."""
    fields = _get_fields(form)
    label_to_id: dict[str, str] = {}
    for f in fields:
        label = f.get("label", "").strip().lower()
        fid = f.get("id", "")
        if label:
            label_to_id[label] = fid
        if fid:
            label_to_id[fid.lower()] = fid

    if filename.lower().endswith((".xlsx", ".xls")):
        raw_rows = _parse_xlsx(raw_bytes)
    else:
        raw_rows = _parse_csv(raw_bytes)

    if not raw_rows:
        return {"fields": fields, "rows": [], "errors": [{"row": 0, "message": "File is empty"}]}

    headers = [str(h).strip() for h in raw_rows[0]]
    data_rows = raw_rows[1:]

    if data_rows and _is_hint_row(data_rows[0], headers):
        data_rows = data_rows[1:]

    parsed_rows: list[dict] = []
    errors: list[dict] = []

    for row_idx, row in enumerate(data_rows, start=3):
        if all(str(v).strip() == "" for v in row):
            continue

        row_dict: dict[str, Any] = {"_row": row_idx}

        for col_idx, header in enumerate(headers):
            value = row[col_idx] if col_idx < len(row) else ""
            value = str(value).strip() if value is not None else ""

            if header.lower() == "asset_id":
                row_dict["asset_id"] = value
            else:
                field_id = label_to_id.get(header.lower())
                if field_id:
                    row_dict[field_id] = value
                else:
                    row_dict[header] = value

        for f in fields:
            if f.get("required") and not row_dict.get(f["id"], "").strip():
                errors.append({
                    "row": row_idx,
                    "message": f"Row {row_idx}: required field '{f.get('label', f['id'])}' is empty",
                })

        parsed_rows.append(row_dict)

    return {
        "fields": [
            {"id": f.get("id"), "label": f.get("label", f.get("id")), "type": f.get("type", "text")}
            for f in fields
        ],
        "rows": parsed_rows,
        "errors": errors,
    }


def _is_hint_row(row: list, headers: list) -> bool:
    hint_keywords = {"text", "number", "date", "select", "options:", "required", "integer", "boolean"}
    for val in row:
        if any(kw in str(val).lower() for kw in hint_keywords):
            return True
    return False


def _parse_csv(raw_bytes: bytes) -> list[list]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw_bytes.decode(encoding)
            reader = csv.reader(io.StringIO(text))
            return [row for row in reader]
        except UnicodeDecodeError:
            continue
    raise ValueError("Unable to decode CSV file. Please use UTF-8 encoding.")


def _parse_xlsx(raw_bytes: bytes) -> list[list]:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel parsing")

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def commit_bulk_rows(
    db: Session,
    form: Form,
    rows: list[dict],
    user_email: str = "system",
) -> dict[str, Any]:
    """Create entity records for each validated row.

    Uses the form schema for required-field validation (not entity_fields)
    to avoid false negatives from field-name mismatches.

    Returns: { "created": N, "errors": [...] }
    """
    from app.models.audit import EntityAuditLog
    from app.services.projection_service import sync_projection

    fields = _get_fields(form)
    field_ids = {f["id"] for f in fields}

    required_field_ids: list[tuple[str, str]] = [
        (f["id"], f.get("label", f["id"]))
        for f in fields
        if f.get("required") and f["id"] != "status"
    ]

    # Entity must exist — enforced strictly
    entity_def = (
        db.query(EntityDefinition)
        .filter(EntityDefinition.name == form.entity_name)
        .first()
    )
    if not entity_def:
        return {
            "created": 0,
            "errors": [{"row": 0, "message": f"Entity '{form.entity_name}' not found"}],
        }

    from .entity_service import _inject_system_fields
    system_fields = [f for f in entity_def.fields if f.is_system]

    created = 0
    errors: list[dict] = []

    for row in rows:
        row_num = row.get("_row", "?")
        asset_id = str(row.get("asset_id", "")).strip() or "N/A"

        data: dict[str, Any] = {}
        record_status = "OPEN"

        for key, value in row.items():
            if key.startswith("_") or key == "asset_id":
                continue
            str_val = str(value).strip() if value is not None else ""
            if key == "status":
                if str_val:
                    record_status = str_val
                continue
            if key in field_ids:
                data[key] = str_val if str_val != "" else None

        # Validate required fields using form schema
        missing_labels = [
            label for fid, label in required_field_ids
            if not data.get(fid)
        ]
        if missing_labels:
            errors.append({
                "row": row_num,
                "message": f"Missing required fields: {missing_labels}",
            })
            continue

        # Inject system fields
        data = _inject_system_fields(data, system_fields)

        try:
            record = EntityRecord(
                entity_id=entity_def.id,
                asset_id=asset_id,
                timestamp=datetime.now(timezone.utc),
                status=record_status,
                data_json=data,
                created_by=user_email,
            )
            db.add(record)
            db.flush()

            db.add(EntityAuditLog(
                entity_record_id=record.id,
                old_data=None,
                new_data=data,
                changed_by=user_email,
            ))

            sync_projection(db, form.entity_name, record)
            db.commit()
            created += 1
        except Exception as exc:
            db.rollback()
            errors.append({"row": row_num, "message": str(exc)})

    return {"created": created, "errors": errors}